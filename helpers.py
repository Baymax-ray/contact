"""
helper functions that support body

# Jiarui Fang (jf659@cornell.edu)
# Feb 12, 2022
"""
import os
import cv2
import bisect
from object_ import *
from random import sample
from constant import *
import numpy as np
import openpyxl


def secondLorS(x,t):
    """
    if t==true, return the second largest number in a list
    if t==false, return the second smallest number in a list
    without changing the original list

    Parameter x: the list
    Precondition: x is a list

    Parameter t: true or false
    Precondition: t is a boolean
    """
    y=copy.deepcopy(x)
    y.sort()
    if t==True:
        return y[-2]
    else:
        return y[1]
    

def read_txt(f_name,x0,y0,x1,y1,time1,time2):
    """
    Return information about the time, the location (x,y), and the neuron activity of the rat from an text

    Note that the data should be seperated by ","

    Parameter f_name: The name of that text
    Precondition: NONE (name can be anything)

    Parameter x0: The left edge of cropped frame
    Precondition: x0 is a number

    Parameter y0: The top edge of cropped frame
    Precondition: y0 is a number

    Parameter x1: The right edge of cropped frame
    Precondition: x1 is a number

    Parameter y1: The lower edge of cropped frame
    Precondition: y1 is a number
    """
    f=open(f_name)
    lines = f.readlines()
    number=len(lines)
    time=[]
    x_l=[]
    y_l=[]
    n_l=[]
    i=0
    while i < number:
        #print(i)
        line=lines[i]
        a=line.find(',')
        b=line.find(',',a+1)
        c=line.find(',',b+1)
        try:
            t=float(line[:a])
            if t>=time1 and t<=time2:
                x=float(line[a+1:b])
                y=float(line[b+1:c])
                if x>x0 and x<x1 and y>y0 and y<y1:
                    x_l.append(x-x0)
                    y_l.append(y-y0)
                    time.append(t)
                    if line.find('\n')>0:
                        n_l.append(float(line[c+1:-1]))
                    else:
                        n_l.append(float(line[c+1:]))
            i=i+1
        except ValueError:
            i=i+1
    assert len(time)==len(x_l)==len(y_l)==len(n_l),"inputs have different length"
    f.close()
    return [time,x_l,y_l,n_l]


def random_list(num,size):
    """
    return a list of num random points inside size

    Parameter num: The amount of points to generate
    Precondition: num is a positive number smaller than size[0]*size[1]

    Parameter size: The size of cropped frame that these points should placed in
    Precondition: size is a list of [width,height]
    """
    x=sample(range(0,size[0]),num)
    y=sample(range(0,size[1]),num)
    result=[]
    for i in range(num):
        result.append([x[i],y[i]])
    #print(str(result))
    return result


def light_helper(points,sta,frame):
    """
    quickly check if the light condition changes
    return True if it changes

    Parameter points: a list of points
    Precondition: points is a list of [x,y]

    Parameter sta: The threshold
    Precondition: sta is a number

    Parameter frame: The frame to check
    Precondition: frame is a picture
    """
    result=True
    l=0
    for p in points:
        l=frame[p[1],p[0]]+l
    if abs(l/len(points)-sta)<0.1*sta:
        result=False
    #print("light!-"+str(l/len(points)))
    return result


def deeperorbrighter(image,x,y,sta):
    """
    Return true if the circle is solid and return false otherwise

    Parameter image: the frame to use
    Precondition: image is a picture

    Parameter x: the x-axis of the circle center
    Precondition: x is int and x>0 and x<wid-5

    Parameter y: the y-axis of the circle center
    Precondition: y is int and y>0 and y<height-5

    Parameter sta: the average brightness of the picture
    Precondition: sta is a number
    """
    # 颠倒xy，因为image[高，宽]
    q=int(x)
    x=int(y)
    y=int(q)
    sum=[]
    shape=image.shape
    s=False
    if x+(DB/2)>=shape[0] or y+(DB/2)>=shape[1] or x-(DB/2)<=0 or y-(DB/2)<=0:
        return s
    c=0
    for t in range(DB):
        for u in range(DB):
            sum.append(int(image[int(x+t-((DB-1)/2)),int(y+u-((DB-1)/2))]))
            #if abs(int(image[int(x+t-((DB-1)/2)),int(y+u-((DB-1)/2))])-sta)>0.2*sta:
                #c=c+1
    med=np.median(sum)
    #if c>=DB*0.7:
    if sta<10: #when it is really dark, the medium is lighter
        if (med-sta)>0.2*sta:
            s=True
    else: #when it is light, the medium is darker
        if (med-sta)<-0.2*sta:
            s=True
    #print (sum)
    #print ("for"+str(x)+"  "+str(y)+" "+str(s))
    return s


def deeperorbrighter_helper(image):
    """
    return the average brightness of the picture
    the larger, the brighter

    Parameter image: the frame to use
    Precondition: image is a picture
    """
    shape=image.shape
    sum=[]
    #s=0
    for y in range(0,shape[0]):
        for x in range (0,shape[1]):
            sum.append(image[y,x])
            #s=s+image[y,x]
            #if x%100==0:
                #print (image[y,x,0]+image[y,x,1]+image[y,x,2])
    #print (s/(shape[0]*shape[1]))
    return np.median(sum)


def circle_detect(image,sta,current,potential):
    """
    Return the circle's center and radius in a list [x,y,r]
    return [0,0,0] if no circle detected

    Parameter image: the frame to use
    Precondition: image is a picture

    Parameter sta: the average brightness of the picture
    Precondition: sta is a number

    Parameter curent: the current object to draw
    Precondition: current is a Currentobject

    Parameter potential: the list of potential objects to draw
    Precondition: potential is a Potentialobject
    """
    # 增加亮度
    # image=increase_brightness(image)
    #img = cv2.adaptiveThreshold(gray,255,cv2.ADAPTIVE_THRESH_MEAN_C,cv2.THRESH_BINARY,15,2)
    # 增加对比度
    #img = cv2.convertScaleAbs(gray,alpha=1.3,beta=0)
    # 进行中值滤波
    img = cv2.medianBlur(image, 5)
    # 霍夫变换圆检测
    circles = cv2.HoughCircles(img, cv2.HOUGH_GRADIENT, 1, 30, param1=100, param2=sens, minRadius=10, maxRadius=30)
    x=0
    y=0
    r=0
    temp=[]
    if not circles is None:
        #print('find')
        for circle in circles[0]:
            # 半径
            r = int(circle[2])
            x = int(circle[0])
            y = int(circle[1])
            # 找出代表medium的圆
            if r>=10 and r<=25 and deeperorbrighter(image,x,y,sta)==True:
                temp.append([x,y,r])
                
    ok=False
    if len(temp)==0:
        potential.pause
        current.noadd()
        current.check_delay()
        current.changestate(False)
    else:
        current.noclear()
        for t in temp:
            if current.check_state(t[0],t[1],t[2])==True:
                ok=True
                #print("continue")
        if ok==True:
        #  this means the current object is still there
            current.changestate(True)
            potential.pause()
        else:
            for t in temp:
                potential.add(t)
            p=potential.check()
            if p!=False:
                current.update(p)
                potential.clear()
    return(current.getlocation()+[current.getradius()])


def contact_helper(current,points):
    """
    return the point that contacts the cup, return -1 otherwise

    Parameter current: the cup
    Precondition: current is [x,y,r]

    Parameter points: infomation of points
    Precondition: points is a list of [x,y,time]
    """
    leng=len(points)
    stat=-1
    dis_list=[]
    time=[]
    for point in points:
        dis=((point[0]-current[0])**2+(point[1]-current[1])**2)**0.5
        dis_list.append(dis)
        time.append(point[2])
        if DISTANCE[0]<=dis<=DISTANCE[1]:
            if stat==-1:
                stat=point
    if len(time)<=1:
        return -1
    model = np.polyfit(time, dis_list, 1)
    if  -100>model[0]>-0.1: #the rat need to go toward the object, but it should not "fly" to the object
        stat=-1
    #print("speed: "+str(model[0]))
    return stat


def file_modify(lines,deles,adds,interval,path):
    """
    modify the txt

    Parameter lines: previous info in the txt
    Precondition: lines is a list, lines[i] is ith line of the txt

    Parameter deles: the frames to delete
    Precondition: deles is a list of numbers

    Parameter adds: the frames with info to add
    Precondition: adds is a list of the info to be added

    Parameter interval: the interval between two frames in sec
    Precondition: interval is a number

    Parameter path: the path of this folder
    Precondition: path is a string
    """
    number=len(lines)
    frame=[]
    i=6
    while i < number:
        #print(i)
        line=lines[i]
        a=line.find(',')
        b=line.find(',',a+1)
        frame.append(int(line[a+1:b]))
        i=i+1
    for add in adds:
        pos=bisect.bisect_left(frame,add[0])+6 #there are 6 lines before the first line of data
        frame.insert(pos-6,add[0])
        time_towrite=str(pos//3600)+":"+str((pos%3600)//60)+":"+str(pos%60)
        lines.insert(time_towrite,str(round((add[0]-L)*interval,2))+","+str(add[0])+","+str(tuple(add[1]))+","+str(add[2])+"\n")
    ddd=[]
    for dele in deles:
        pos=bisect.bisect_left(frame,dele)+6
        ddd.append(pos)
    fil=open(path+"\\"+"data file.txt",'w')
    number2=len(lines)
    #print(lines)
    #print(ddd)
    for i in range(number2):
        if (i in ddd)==False:
            #print(lines[i])
            fil.write(lines[i])
    fil.close()


def contact_decider(point,last):
    """
    return true if the contact should be cout: its time or location must be very different from the last contact

    Parameter point: the point of contact
    Precondition: point is a list of x, y, time

    Parameter last: the last point
    Precondition: last is a list of x, y, time

    """
    result=False
    if point!=-1:
        if point[2]-last[2]>GAP:
            result=True
        #elif (point[0]-last[0])**2+(point[1]-last[1])**2>(2*RAD)**2:
            #print(str((point[0]-last[0])**2+(point[1]-last[1])**2))
            #result=True

    return result

def conformcount(time,count):
    """
    return the list of all starts of trials if we have the correct amount of trials, otherwise return the amount of timestamps missing

    Parameter time: the timestamps
    Precondition: time is a list of numbers

    Parameter count: the amount of trials
    Precondition: count is a number
    """
    print(str(time[0])+"--this is the first timestamp")
    fisrt=int(input("Is this the start of first trial?(type 1 for yes, 0 for the indicator of start)"))
    print(str(time[-1])+"--this is the last timestamp")
    last=int(input("Is this the start of last trial?(type 1 for yes, 0 for the indicator of end)"))
    if fisrt==1:
        if last==1:
            print("so the first trial starts at "+str(time[0])+" and the last trial starts at "+str(time[-1]))
            if len(time)==count:
                return time
            else:
                return count-len(time)
        else: #last is 0
            print("so the first trial starts at "+str(time[0])+" and the last trial starts at "+str(time[-2]))
            if len(time)-1==count:
                return time[:-1]
            else:
                return count-len(time)+1
    else: #fisrt is 0
        if last==1:
            print("so the first trial starts at "+str(time[1])+" and the last trial starts at "+str(time[-1]))
            if len(time)-1==count:
                return time[1:]
            else:
                return count-len(time)+1
        else: #last is 0
            print("so the first trial starts at "+str(time[1])+" and the last trial starts at "+str(time[-2]))
            if len(time)-2==count:
                return time[1:-1]
            else:
                return count-len(time)+2

def gettimestamps(path,name):
    """
    return timestamps in a list of number

    Parameter path: the path of this folder
    Precondition: path is a string

    Parameter name: the name of the video
    Precondition: name is a string
    """
    f=open(path+"\\"+name+".txt",'r')
    lines = f.readlines()
    number=len(lines)
    time=[]
    splitpoints=0
    i=0
    while i < number:
        #print(i)
        line=lines[i]
        a=line.find(',')
        b=line.rfind(',')
        de=line[-1]=='\n' and line[-2]=='1'
        if de or line[-1]=='1':
            time.append(float(line[:a]))
            splitpoints=splitpoints+1
        i=i+1
    f.close()
    #print(time)
    return time


def rat_detecter(frame):
    """
    return the position of the rat --y,x--, which is the midpoint of red and green point
    return None is not found

    Parameter frame: the picture to use
    Precondition: frame is a frame of a video
    """
    red=frame[:,:,2].astype(np.float32) #need to allow negative values
    green=frame[:,:,1].astype(np.float32) #need to allow negative values
    blue=frame[:,:,0].astype(np.float32) #need to allow negative values
    r=red-green-blue
    g=green-red-blue
    a=r.max()
    b=g.max()
    #print(a)
    #print(b)
    if a>S and b>S:
        aa=np.where(r==a)
        bb=np.where(g==b)
        aaa=np.mean(aa,axis=1) #mean of the red
        bbb=np.mean(bb,axis=1) #mean of the green
        # show x and y
        return (aaa[0]+bbb[0])//2,(aaa[1]+bbb[1])//2
    else:
        #write a and b on this frame
        cv2.putText(frame,str(a),(0,50),cv2.FONT_HERSHEY_SIMPLEX,1,(0,255,255),2)
        cv2.putText(frame,str(b),(0,100),cv2.FONT_HERSHEY_SIMPLEX,1,(0,255,255),2)
        return None,None


def prepare(frame):
    """
    return a frame with pre-treatment

    Parameter frame: the frame to be treated
    """
    # 增加对比度
    image = cv2.convertScaleAbs(frame,alpha=1.5,beta=0)
    # 灰度化
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    # 进行中值滤波
    img = cv2.medianBlur(gray, 5)
    return img

def setstandard_helper(path,cupsinwhite,cupsinblack,sequence):
    """
    return the standard brightness of cups in different boxes

    Parameter path: the path of this folder
    Precondition: path is a string

    Parameter cupsinwhite: the list of cups in white box
    Precondition: cupsinwhite is a list of numbers [[x,y],[x,y]]

    Parameter cupsinblack: the list of cups in black box
    Precondition: cupsinblack is a list of numbers [[x,y],[x,y]]

    Parameter sequence: the sequence of the boxes
    Precondition: sequence is 1 or 2; 1 for white box first, 2 for black box first
    """
    white = cv2.imread(path+"\\"+"white.jpg")
    #white2 = cv2.imread(path+"\\"+"white2.jpg")
    black = cv2.imread(path+"\\"+"black.jpg")
    #black2 = cv2.imread(path+"\\"+"black2.jpg")
    #prepare the pictures
    white1=prepare(white)
    #white2=prepare(white2)
    black1=prepare(black)
    #black2=prepare(black2)
    
    #get the brightness of the cups in the white box
    a=int(cupsinwhite[0][1])
    b=int(cupsinwhite[0][0])
    c=int(cupsinwhite[1][1])
    d=int(cupsinwhite[1][0])
    white1brightness=white1[int(a-RAD/1.5):int(a+RAD/1.5),int(b-RAD/1.5):int(b+RAD/1.5)].mean()+\
            white1[int(c-RAD/1.5):int(c+RAD/1.5),int(d-RAD/1.5):int(d+RAD/1.5)].mean()
    #white2brightness=white2[int(a-RAD/1.5):int(a+RAD/1.5),int(b-RAD/1.5):int(b+RAD/1.5)].mean()+\
            #white2[int(c-RAD/1.5):int(c+RAD/1.5),int(d-RAD/1.5):int(d+RAD/1.5)].mean()
    #get the brightness of the cups in the black box
    a=int(cupsinblack[0][1])
    b=int(cupsinblack[0][0])
    c=int(cupsinblack[1][1])
    d=int(cupsinblack[1][0])
    black1brightness=black1[int(a-RAD/1.5):int(a+RAD/1.5),int(b-RAD/1.5):int(b+RAD/1.5)].mean()+\
            black1[int(c-RAD/1.5):int(c+RAD/1.5),int(d-RAD/1.5):int(d+RAD/1.5)].mean()
    #black2brightness=black2[int(a-RAD/1.5):int(a+RAD/1.5),int(b-RAD/1.5):int(b+RAD/1.5)].mean()+\
            #black2[int(c-RAD/1.5):int(c+RAD/1.5),int(d-RAD/1.5):int(d+RAD/1.5)].mean()
    print(white1brightness,black1brightness)
    #print(white1brightness,white2brightness,black1brightness,black2brightness)
    #get the standard brightness by divided by 2
    #return white1brightness//2,white2brightness//2,black1brightness//2,black2brightness//2
    return white1brightness//2,black1brightness//2

def cup_detecter(frame,cups,rat,box):
    """
    return 1 if the cup1 is possible to be the cup in the box, 2 if the cup2 is possible
    return None is not found

    Parameter frame: the picture to use
    Precondition: frame is a frame of a video

    Parameter cups: the potential position of cups in the box
    Precondition: cups is a list of coordinates [[x,y],[x,y],...]

    Parameter rat: the position of the rat
    Precondition: rat is a list of numbers [x,y]

    Parameter box: the color of the boxes
    Precondition: box is 0 or 1; 0 for white box, 1 for black box
    """
    #prepare the frame
    img=prepare(frame)
    number_cups=len(cups)
    #print("number of cups:",number_cups)
    cups_brightness=[]
    
    for i in range(number_cups):
        a=int(cups[i][1])
        b=int(cups[i][0])
        cups_brightness.append(img[int(a-RAD/1.5):int(a+RAD/1.5),int(b-RAD/1.5):int(b+RAD/1.5)].mean())
    if rat[0]!=None and rat[1]!=None and rat[0]-RAD<=a<=rat[0]+RAD and rat[1]-RAD<=b<=rat[1]+RAD:
        return None
    else:
        if box==0: #white box black cups
            #if one element is V smaller than all the others, it is the cup
            result=cups_brightness.index(min(cups_brightness))
            cups_brightness.sort()
            if cups_brightness[0]+V<cups_brightness[1]:
                return result+1
        elif box==1: #black box white cups
            #if one element is V larger than all the others, it is the cup
            result=cups_brightness.index(max(cups_brightness))
            cups_brightness.sort()
            if cups_brightness[-1]>45 and cups_brightness[-1]-V>cups_brightness[-2]:
                return result+1
    return None


def writedata(Trials,sequence,name,path2,right_timestamp,timesplit,cupsinbox,fps):
    """
    write the data of the experiment into the file

    Parameter Trials: the list of trials
    Precondition: Trials is a list of objects of Trial class

    Parameter timesplit: the time when a box is finished
    Precondition: timesplit is a list of number

    Parameter sequence: the sequence of the trials
    Precondition: sequence is 1 or 2; 1 for white box first, 2 for black box first

    Parameter name: the name of the experiment
    Precondition: name is a string

    Parameter path2: the path to the file
    Precondition: path2 is a string

    Parameter right_timestamps: the timestamps of the start of trials
    Precondition: right_timestamps is a list of numbers

    Parameter cupsinbox: the position of cups in a box
    Precondition: cupsinbox is a list of a list of two coordinates-[[[x,y],[x,y],...],[[x,y],[x,y],...],...]
    
    Parameter fps: the frame per second of the video
    Precondition: fps is a number
    """
    print("writing data")
    boxcount=len(cupsinbox)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # titles
    sheet.title = "data"
    sheet.cell(row=1, column=1).value = "Trial"
    sheet.cell(row=1, column=2).value = "Start time"
    sheet.cell(row=1, column=3).value = "End time"
    sheet.cell(row=1, column=4).value = "B/W"
    maxcup=0
    for i in range(boxcount):
        if len(cupsinbox[i])>maxcup:
            maxcup=len(cupsinbox[i])
    for i in range(maxcup):
        sheet.cell(row=1, column=5+i).value = "Cup"+str(i+1)
    sheet.cell(row=1, column=5+maxcup).value = "Comment"
    sheet.cell(row=1, column=6+maxcup).value = "fps"
    
    # data
    sheet.cell(row=2, column=6+maxcup).value = fps
    for j in range(len(Trials)):
        n=bisect.bisect(timesplit,right_timestamp[j])
        boxcolor=0 if sequence%2!=n%2 else 1
        trial=Trials[j].gettrial()
        sheet.cell(row=j+2, column=1).value = trial
        sheet.cell(row=j+2, column=2).value = secondstohms(right_timestamp[trial-1])
        if j+1==len(Trials): #the last trial
            sheet.cell(row=j+2, column=3).value = None
        else:
            m=bisect.bisect(timesplit,right_timestamp[j+1])
            if n==m: #in the same trial
                sheet.cell(row=j+2, column=3).value = secondstohms(right_timestamp[trial])
            else: #the last trial of (n+1)th box
                sheet.cell(row=j+2, column=3).value = secondstohms(timesplit[n])
        sheet.cell(row=j+2, column=4).value = "W" if boxcolor==0 else "B"
        cup=Trials[j].getobject()
        if cup==None:
            sheet.cell(row=j+2, column=5+maxcup).value = "I do not know which cup to use!"
        else:
            sheet.cell(row=j+2, column=4+cup).value = str(cupsinbox[n][cup-1])

    #save
    workbook.save(path2+"\\"+name+"_data"+".xlsx")
    print("Data saved")

def secondstohms(seconds):
    """
    convert seconds to hh:mm:ss

    Parameter seconds: the number of seconds
    Precondition: seconds is a number
    """
    seconds=float(seconds)
    m, s = divmod(seconds, 60)
    h, m = divmod(m, 60)
    ss=round(s-int(s),2)
    return "%d:%02d:%02d" % (h, m, s)+"."+str(int(ss*100))
    
def hmstoseconds(time):
    """
    convert hh:mm:ss to seconds

    Parameter time: the time in hh:mm:ss
    Precondition: time is a string
    """
    time=str(time)
    h, m, s = time.split(':')
    return int(h) * 3600 + int(m) * 60 + float(s)

def convertformat(sheet,fps=0):
    """
    convert the sheet to required format
    Will try to convert the first line, but will do nothing if it is the title

    Parameter sheet: the sheet to be converted
    Precondition: sheet is a sheet in excel

    Parameter fps: the frame per second of the video
    Precondition: fps is a number
    """
    max=sheet.max_row
    #print(max)
    if fps==0:
        try:
            value=sheet.cell(row=1, column=1).value
            a=value.find(':')
            b=value.find(':',a+1)
            c=value.find('.')
            hour=value[:a]
            min=value[a+1:b]
            sec=value[b+1:c]
            misec=value[c+1:]
            sheet.cell(row=1, column=2).value=int(hour)
            sheet.cell(row=1, column=3).value=int(min)
            sheet.cell(row=1, column=4).value=int(sec)
            sheet.cell(row=1, column=5).value=int(misec)
            total=int(hour)*3600+int(min)*60+int(sec)+float(misec)/10
            sheet.cell(row=1, column=6).value=total
        except:
            pass
        for x in range(2,max+1):
            value=sheet.cell(row=x, column=1).value
            a=value.find(':')
            b=value.find(':',a+1)
            c=value.find('.')
            hour=value[:a]
            min=value[a+1:b]
            sec=value[b+1:c]
            smisec=value[b+1:]
            misec=1000*(float(smisec)-int(sec))
            total=int(hour)*3600+int(min)*60+int(sec)+float(misec)/1000
            sheet.cell(row=x, column=2).value=int(hour)
            sheet.cell(row=x, column=3).value=int(min)
            sheet.cell(row=x, column=4).value=int(sec)
            sheet.cell(row=x, column=5).value=round(misec,2)
            sheet.cell(row=x, column=6).value=total
    else:
        try:
            frame=int(float(sheet.cell(row=1, column=1).value))
            seconds=round(frame*(1/fps),2)
            #convert seconds to h:m:s.ms
            hour=int(seconds/3600)
            min=int((seconds-hour*3600)/60)
            sec=int(seconds-hour*3600-min*60)
            secs=round(seconds-hour*3600-min*60,2)
            misec=round((secs-sec)*1000,2)
            together=str(hour)+':'+str(min)+':'+str(secs)
            sheet.cell(row=1, column=7).value=seconds
            sheet.cell(row=1, column=2).value=together
            sheet.cell(row=1, column=3).value=int(hour)
            sheet.cell(row=1, column=4).value=int(min)
            sheet.cell(row=1, column=5).value=int(sec)
            sheet.cell(row=1, column=6).value=int(misec)
        except:
            pass
        for x in range(2,max+1):
            if type(sheet.cell(row=x, column=1).value)==int or type(sheet.cell(row=x, column=1).value)==float: #this is a data
                frame=int(sheet.cell(row=x, column=1).value)
                seconds=round(frame*(1/fps),2)
                #convert seconds to h:m:s.ms
                hour=int(seconds/3600)
                min=int((seconds-hour*3600)/60)
                sec=int(seconds-hour*3600-min*60)
                secs=round(seconds-hour*3600-min*60,2)
                misec=round((secs-sec)*1000,2)
                together=str(hour)+':'+str(min)+':'+str(secs)
                sheet.cell(row=x, column=7).value=seconds
                sheet.cell(row=x, column=2).value=together
                sheet.cell(row=x, column=3).value=int(hour)
                sheet.cell(row=x, column=4).value=int(min)
                sheet.cell(row=x, column=5).value=int(sec)
                sheet.cell(row=x, column=6).value=int(misec)
            else: #this is a error message
                pass

def remove_pre_data(path,i):
    """
    remove the pre pictures in the 'trials' folder

    Parameter path: the path of the folder
    Precondition: path is a string

    Parameter i: the # of a trial
    Precondition: i is a number
    """
    #remove the pre data in the folder
    for relpath, dirs, files in os.walk(path):
        for file in files:
            if str(i)+"_" in file:
                os.remove(relpath+"/"+file)